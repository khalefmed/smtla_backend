from datetime import timedelta
from rest_framework import generics
import csv, io
import os
from django.conf import settings
from django.http import HttpResponse

from .models import *
from .serializers import *

from django.db import transaction as db_transaction
from django.db.models.functions import TruncMonth

import pandas as pd
from django.db import transaction

from rest_framework.decorators import api_view, permission_classes
from django.utils.crypto import get_random_string

from django.http import HttpRequest
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView, TokenVerifyView
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from rest_framework_simplejwt.tokens import RefreshToken, TokenError
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import check_password, make_password

from rest_framework.pagination import PageNumberPagination

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json

from django.db.models import Q, Avg, Sum, Count, F
from django.core.files.storage import default_storage
from django.db.models import Sum, Count, Q, FloatField
from django.db.models.functions import Coalesce
from datetime import datetime, timedelta
import calendar
from rest_framework import status
from openpyxl import load_workbook

import tempfile
import subprocess
from django.http import FileResponse
from django.shortcuts import get_object_or_404

from django.db.models.functions import TruncDate
from collections import defaultdict


# ==================== DASHBOARD ====================

class DashboardStatsView(APIView):
    """Statistiques du tableau de bord enrichies - Spec 8"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        role = user.type
        stats = {}

        print(f"Utilisateur: {user.username}, Rôle: {role}")

        # 1. Statistiques Rotations & Stocks
        if role in ['agent_port', 'directeur_operations', 'directeur_general']:
            stats['rotations'] = {
                'total_entrantes': RotationEntrante.objects.count(),
                'total_sortantes': RotationSortante.objects.count(),
            }

            print(stats['rotations'])

            stocks = []
            for client in Client.objects.all():
                client_stock = {'client': client.nom, 'types': []}
                for type_mat in TypeMateriel.objects.all():
                    entrees = RotationEntrante.objects.filter(client=client, type_materiel=type_mat).aggregate(Sum('quantite'))['quantite__sum'] or 0
                    sorties = RotationSortante.objects.filter(client=client, type_materiel=type_mat).aggregate(Sum('quantite'))['quantite__sum'] or 0
                    disponible = entrees - sorties
                    if disponible > 0:
                        client_stock['types'].append({
                            'type_materiel': type_mat.nom,
                            'quantite_disponible': disponible
                        })
                if client_stock['types']:
                    stocks.append(client_stock)
            stats['stocks_par_client'] = stocks

        # 2. Statistiques Clients
        if role in ['agent_port', 'comptable', 'directeur_operations', 'directeur_general']:
            stats['clients'] = { 'total': Client.objects.count() }

        # 3. Expression de Besoin (Compteurs + Liste en attente)
        if role in ['assistant', 'comptable', 'directeur_operations', 'directeur_general']:
            eb_attente_query = ExpressionBesoin.objects.filter(status='attente')
            stats['expressions_besoin'] = {
                'en_attente': eb_attente_query.count(),
                'en_cours': ExpressionBesoin.objects.filter(status='en_cours').count(),
                'validees': ExpressionBesoin.objects.filter(status='valide').count(),
                'rejetees': ExpressionBesoin.objects.filter(status='rejete').count(),
                # On ajoute la liste détaillée ici
                'liste_en_attente': ExpressionBesoinSerializer(eb_attente_query, many=True).data
            }

        # 4. Note de Frais
        if role in ['assistant', 'comptable', 'directeur_operations', 'directeur_general']:
            stats['notes_frais'] = {
                'en_attente': NoteDeFrais.objects.filter(status='attente').count(),
                'validees': NoteDeFrais.objects.filter(status='valide').count(),
                'total_montant': NoteDeFrais.objects.filter(status='valide').aggregate(Sum('items__montant'))['items__montant__sum'] or 0
            }

        # 5. Devis & Factures (Compteurs + Liste Devis en attente)
        if role in ['comptable', 'directeur_operations', 'directeur_general']:
            devis_attente_query = Devis.objects.filter(status='attente')
            stats['devis'] = {
                'en_attente': devis_attente_query.count(),
                'valides': Devis.objects.filter(status='valide').count(),
                'total': Devis.objects.count(),
                'somme_totale': sum(d.montant_total for d in Devis.objects.all()),
                # On ajoute la liste détaillée ici
                'liste_en_attente': DevisSerializer(devis_attente_query, many=True).data
            }

            factures_query = Facture.objects.all()
            if role != 'directeur_general':
                factures_query = factures_query.filter(est_privee=False)

            stats['factures'] = {
                'en_attente': factures_query.filter(status='attente').count(),
                'validees': factures_query.filter(status='valide').count(),
                'total': factures_query.count(),
                'somme_totale': sum(f.montant_total for f in factures_query)
            }

        # 6. Bons de Commande
        if role in ['comptable', 'directeur_operations', 'directeur_general']:
            stats['bons_commande'] = {
                'en_attente': BonCommande.objects.filter(status='attente').count(),
                'valides': BonCommande.objects.filter(status='valide').count(),
                'total': BonCommande.objects.count(),
            }

        return Response(stats)



class StockStatusView(APIView):
    """
    Retourne l'état des stocks actuel par client et par type de matériel.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        stocks = []
        clients = Client.objects.all()
        types_mat = TypeMateriel.objects.all()

        for client in clients:
            client_stock = {
                'client': client.nom,
                'types': []
            }

            for type_mat in types_mat:
                # Calcul des entrées
                total_entrees = RotationEntrante.objects.filter(
                    client=client,
                    type_materiel=type_mat
                ).aggregate(Sum('quantite'))['quantite__sum'] or 0

                # Calcul des sorties
                total_sorties = RotationSortante.objects.filter(
                    client=client,
                    type_materiel=type_mat
                ).aggregate(Sum('quantite'))['quantite__sum'] or 0

                disponible = total_entrees - total_sorties

                # On n'ajoute que si le client a déjà eu ce matériel en stock
                if total_entrees > 0:
                    client_stock['types'].append({
                        'type_materiel': type_mat.nom,
                        'quantite_disponible': disponible,
                        'total_entrees': total_entrees,
                        'total_sorties': total_sorties
                    })

            if client_stock['types']:
                stocks.append(client_stock)

        return Response(stocks)


# ==================== TYPE MATERIEL (Spec 5) ====================

class TypeMaterielListCreateView(generics.ListCreateAPIView):
    """Liste tous les types de matériel ou crée un nouveau - Spec 5"""
    queryset = TypeMateriel.objects.all().order_by('nom')
    serializer_class = TypeMaterielSerializer
    permission_classes = [IsAuthenticated]


class TypeMaterielRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un type de matériel - Spec 5"""
    queryset = TypeMateriel.objects.all()
    serializer_class = TypeMaterielSerializer
    permission_classes = [IsAuthenticated]


class TypeMaterielRechercheView(generics.ListAPIView):
    """Recherche de types de matériel par nom - Spec 5"""
    serializer_class = TypeMaterielSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        query = self.request.query_params.get('q', '')
        return TypeMateriel.objects.filter(
            Q(nom__icontains=query) | Q(description__icontains=query)
        ).order_by('nom')


# ==================== ROTATION ====================

class RotationListCreateView(generics.ListCreateAPIView):
    """Liste toutes les rotations ou crée une nouvelle"""
    queryset = Rotation.objects.all().order_by('-date_rotation')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return RotationCreateSerializer
        return RotationSerializer


class RotationRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime une rotation"""
    queryset = Rotation.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return RotationCreateSerializer
        return RotationSerializer


class RotationParTypeView(generics.ListAPIView):
    """Liste les rotations filtrées par type (entrée/sortie)"""
    serializer_class = RotationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        type_rotation = self.request.query_params.get('type')
        queryset = Rotation.objects.all().order_by('-date_rotation')

        if type_rotation:
            queryset = queryset.filter(type=type_rotation)

        return queryset



class TerminerToutesLesRotationsView(APIView):
    """
    API pour passer TOUTES les rotations entrantes et sortantes à 'termine'
    """
    def post(self, request):
        try:
            # Mise à jour massive des entrantes
            nb_entrantes = RotationEntrante.objects.filter(status='en_cours').update(status='termine')

            # Mise à jour massive des sortantes
            nb_sortantes = RotationSortante.objects.filter(status='en_cours').update(status='termine')

            return Response({
                "message": "Toutes les rotations ont été clôturées.",
                "details": {
                    "entrantes_cloturees": nb_entrantes,
                    "sortantes_cloturees": nb_sortantes
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ==================== ROTATIONS ENTRANTES (Spec 6) ====================

class RotationEntranteListCreateView(generics.ListCreateAPIView):
    """Liste les rotations entrantes EN COURS ou crée une nouvelle - Spec 6"""
    queryset = RotationEntrante.objects.all().order_by('-date_arrivee')
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # On filtre pour ne retourner que les rotations "en cours"
        return RotationEntrante.objects.filter(status='en_cours').order_by('-date_arrivee')

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return RotationEntranteCreateSerializer
        return RotationEntranteSerializer


class RotationEntranteRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime une rotation entrante - Spec 6"""
    queryset = RotationEntrante.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return RotationEntranteCreateSerializer
        return RotationEntranteSerializer


class RotationEntranteRapportView(APIView):
    """Génère un rapport journalier des rotations entrantes - Spec 6"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')

        queryset = RotationEntrante.objects.all()

        if date_debut:
            queryset = queryset.filter(date_arrivee__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(date_arrivee__lte=date_fin)

        # Grouper par client et type
        rapport = {}
        for rotation in queryset:
            client_nom = rotation.client.nom
            type_materiel = rotation.type_materiel.nom

            if client_nom not in rapport:
                rapport[client_nom] = {}

            if type_materiel not in rapport[client_nom]:
                rapport[client_nom][type_materiel] = {
                    'quantite_totale': 0,
                    'nombre_rotations': 0,
                    'rotations': []
                }

            rapport[client_nom][type_materiel]['quantite_totale'] += rotation.quantite
            rapport[client_nom][type_materiel]['nombre_rotations'] += 1
            rapport[client_nom][type_materiel]['rotations'].append({
                'id': rotation.id,
                'numero_bordereau': rotation.numero_bordereau,
                'date_arrivee': rotation.date_arrivee,
                'camion': rotation.camion,
                'quantite': rotation.quantite,
                'observation': rotation.observation
            })

        return Response({
            'date_debut': date_debut,
            'date_fin': date_fin,
            'rapport': rapport
        })


# ==================== ROTATIONS SORTANTES (Spec 7) ====================

class RotationSortanteListCreateView(generics.ListCreateAPIView):
    """Liste les rotations sortantes EN COURS ou crée une nouvelle - Spec 7"""
    queryset = RotationSortante.objects.all().order_by('-date_sortie')
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # On filtre pour ne retourner que les rotations "en cours"
        return RotationSortante.objects.filter(status='en_cours').order_by('-date_sortie')

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return RotationSortanteCreateSerializer
        return RotationSortanteSerializer


class RotationSortanteRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime une rotation sortante - Spec 7"""
    queryset = RotationSortante.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return RotationSortanteCreateSerializer
        return RotationSortanteSerializer


class RotationSortanteRapportView(APIView):
    """Génère un rapport de livraison des rotations sortantes - Spec 7"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')

        queryset = RotationSortante.objects.all()

        if date_debut:
            queryset = queryset.filter(date_sortie__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(date_sortie__lte=date_fin)

        # Grouper par client et type
        rapport = {}
        total_rotations = queryset.count()

        for rotation in queryset:
            client_nom = rotation.client.nom
            type_materiel = rotation.type_materiel.nom

            if client_nom not in rapport:
                rapport[client_nom] = {}

            if type_materiel not in rapport[client_nom]:
                rapport[client_nom][type_materiel] = {
                    'quantite_totale': 0,
                    'nombre_rotations': 0,
                    'rotations': []
                }

            rapport[client_nom][type_materiel]['quantite_totale'] += rotation.quantite
            rapport[client_nom][type_materiel]['nombre_rotations'] += 1
            rapport[client_nom][type_materiel]['rotations'].append({
                'id': rotation.id,
                'numero_bordereau': rotation.numero_bordereau,
                'date_sortie': rotation.date_sortie,
                'camion': rotation.camion,
                'quantite': rotation.quantite,
                'observation': rotation.observation
            })

        return Response({
            'date_debut': date_debut,
            'date_fin': date_fin,
            'total_rotations': total_rotations,
            'rapport': rapport
        })


# ==================== CLIENT ====================

class ClientListCreateView(generics.ListCreateAPIView):
    """Liste tous les clients ou crée un nouveau client"""
    queryset = Client.objects.all().order_by('nom')
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated]


class ClientRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un client spécifique"""
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated]


class ClientRechercheView(generics.ListAPIView):
    """Recherche de clients par nom, téléphone ou email"""
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        query = self.request.query_params.get('q', '')
        return Client.objects.filter(
            Q(nom__icontains=query) |
            Q(telephone__icontains=query) |
            Q(email__icontains=query) |
            Q(nif__icontains=query)
        ).order_by('nom')


# ==================== FOURNISSEUR (Spec 9) ====================

class FournisseurListCreateView(generics.ListCreateAPIView):
    """Liste tous les fournisseurs ou crée un nouveau - Spec 9"""
    queryset = Fournisseur.objects.all().order_by('nom')
    serializer_class = FournisseurSerializer
    permission_classes = [IsAuthenticated]


class FournisseurRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un fournisseur - Spec 9"""
    queryset = Fournisseur.objects.all()
    serializer_class = FournisseurSerializer
    permission_classes = [IsAuthenticated]


class FournisseurRechercheView(generics.ListAPIView):
    """Recherche de fournisseurs - Spec 9"""
    serializer_class = FournisseurSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        query = self.request.query_params.get('q', '')
        return Fournisseur.objects.filter(
            Q(nom__icontains=query) |
            Q(raison_sociale__icontains=query) |
            Q(nif__icontains=query) |
            Q(email__icontains=query)
        ).order_by('nom')


# ==================== EXPRESSION DE BESOIN (Spec 1) ====================

class ExpressionBesoinListCreateView(generics.ListCreateAPIView):
    """Liste toutes les expressions de besoin ou crée une nouvelle - Spec 1"""
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return ExpressionBesoinCreateSerializer
        return ExpressionBesoinSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = ExpressionBesoin.objects.all().order_by('-date_creation')

        # Visibilité restreinte (Spec 1)
        if user.type not in ['directeur_operations', 'comptable']:
            # Utilisateur ne voit que ses propres expressions
            queryset = queryset.filter(createur=user)

        return queryset


class ExpressionBesoinRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime une expression de besoin - Spec 1"""
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return ExpressionBesoinCreateSerializer
        return ExpressionBesoinSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = ExpressionBesoin.objects.all()

        # Visibilité restreinte (Spec 1)
        if user.type not in ['directeur_operations', 'comptable']:
            queryset = queryset.filter(createur=user)

        return queryset


class ExpressionBesoinValiderView(APIView):
    """Valide ou rejette une expression de besoin et génère une Note de Frais - Spec 1 & 2"""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        # Vérification des permissions
        if request.user.type not in ['directeur_operations', 'comptable', 'directeur_general']:
            return Response(
                {"error": "Vous n'avez pas la permission de valider"},
                status=status.HTTP_403_FORBIDDEN
            )

        expression = get_object_or_404(ExpressionBesoin, pk=pk)
        nouveau_statut = request.data.get('status')

        if nouveau_statut not in ['valide', 'rejete', 'en_cours']:
            return Response(
                {"error": "Statut invalide"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                # 1. Mise à jour de l'expression de besoin
                expression.status = nouveau_statut
                expression.valideur = request.user
                expression.date_validation = datetime.now()
                expression.save()

                note_frais = None
                # 2. Création automatique de la Note de Frais si validée
                if nouveau_statut == 'valide':
                    # On vérifie si une note n'existe pas déjà pour éviter les doublons
                    if not NoteDeFrais.objects.filter(expression_besoin=expression).exists():
                        note_frais = NoteDeFrais.objects.create(
                            expression_besoin=expression,
                            createur=expression.createur, # Le créateur de l'EB reste le bénéficiaire de la NF
                            status='attente' # La NF commence son propre cycle de validation
                        )

                        # 3. Duplication des items (EB -> NF)
                        items_eb = expression.items.all()
                        items_nf = []

                        for item in items_eb:
                            # Note: On mappe les types de l'EB vers les types de la NF
                            # Si les types sont identiques, c'est direct.
                            items_nf.append(ItemNoteDeFrais(
                                note_de_frais=note_frais,
                                libelle=item.libelle,
                                type=item.type,
                                montant=item.montant
                            ))

                        if items_nf:
                            ItemNoteDeFrais.objects.bulk_create(items_nf)

            # Préparation du message de retour
            msg = f"Expression de besoin {nouveau_statut}e avec succès"
            if note_frais:
                msg += f". Note de frais {note_frais.reference} générée."

            return Response({
                "message": msg,
                "status": expression.status,
                "note_frais_ref": note_frais.reference if note_frais else None
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Erreur lors de la validation : {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ==================== NOTE DE FRAIS (Spec 2) ====================

class NoteDeFraisListCreateView(generics.ListCreateAPIView):
    """Liste toutes les notes de frais ou crée une nouvelle - Spec 2"""
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return NoteDeFraisCreateSerializer
        elif self.request.method == 'GET':
            return NoteFraisDetailSerializer
        return NoteDeFraisSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = NoteDeFrais.objects.all().order_by('-date_creation')

        # Logic for "Comptable": only show validated notes
        if hasattr(user, 'type') and user.type == 'comptable':
            return queryset.filter(status='valide')

        return queryset


class NoteDeFraisRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime une note de frais - Spec 2"""
    queryset = NoteDeFrais.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return NoteDeFraisCreateSerializer
        return NoteDeFraisSerializer


class NoteDeFraisValiderView(APIView):
    """Valide ou rejette une note de frais - Spec 2"""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        note = get_object_or_404(NoteDeFrais, pk=pk)
        nouveau_statut = request.data.get('status')

        if nouveau_statut not in ['valide', 'rejete', 'attente']:
            return Response(
                {"error": "Statut invalide"},
                status=status.HTTP_400_BAD_REQUEST
            )

        note.status = nouveau_statut
        if nouveau_statut in ['valide', 'rejete']:
            note.valideur = request.user
            note.date_validation = datetime.now()
        note.save()

        return Response({
            "message": f"Note {nouveau_statut} avec succès",
            "status": note.status
        })


class NoteDeFraisCreerDepuisExpressionView(APIView):
    """Crée une note de frais à partir d'une expression de besoin - Spec 2"""
    permission_classes = [IsAuthenticated]

    def post(self, request, expression_id):
        expression = get_object_or_404(ExpressionBesoin, pk=expression_id)

        with db_transaction.atomic():
            # Créer la note de frais (le client, navire, etc. sont portés par l'EB)
            note = NoteDeFrais.objects.create(
                expression_besoin=expression,
                createur=request.user
            )

            # Copier les items de l'EB vers la NF pour permettre l'ajustement ultérieur
            for item_eb in expression.items.all():
                ItemNoteDeFrais.objects.create(
                    note_de_frais=note,
                    libelle=item_eb.libelle,
                    type=item_eb.type,
                    montant=item_eb.montant
                )

        return Response(
            {
                "message": "Note de frais générée avec succès",
                "note": NoteDeFraisSerializer(note).data
            },
            status=status.HTTP_201_CREATED
        )


class NoteDeFraisParDeviseView(generics.ListAPIView):
    """Liste les notes de frais filtrées par la devise de l'EB source"""
    serializer_class = NoteDeFraisSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        devise = self.request.query_params.get('devise')
        queryset = NoteDeFrais.objects.all().order_by('-date_creation')

        if devise:
            # Filtrage à travers la relation ForeignKey
            queryset = queryset.filter(expression_besoin__devise=devise)

        return queryset


class NoteDeFraisAjouterItemView(APIView):
    """Ajoute un item à une note de frais existante"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            note = NoteDeFrais.objects.get(pk=pk)
        except NoteDeFrais.DoesNotExist:
            return Response(
                {"detail": "Note de frais introuvable"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = ItemNoteDeFraisSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(note_de_frais=note)
            return Response(
                NoteDeFraisSerializer(note).data,
                status=status.HTTP_201_CREATED
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class NoteDeFraisExportXlsxView(APIView):
    """Export d'une note de frais en XLSX"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            note = NoteDeFrais.objects.get(pk=pk)
        except NoteDeFrais.DoesNotExist:
            return Response({"detail": "Note de frais introuvable"}, status=404)

        template_path = os.path.join(settings.BASE_DIR, "templates", "NOTE-DE-DEPENSES-MRU.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active

        ws["J1"] = note.reference
        ws["J2"] = note.date_creation.date().strftime("%d/%m/%Y")

        start_row = 7
        current_row = start_row
        col_map = {
            "nourriture": "C",
            "hebergement": "D",
            "medicament": "E",
            "carburant": "F",
            "entretien": "G",
            "telecom": "H",
            "avance": "I",
            "divers": "J",
        }
        for item in note.items.all():
            ws[f"A{current_row}"] = item.libelle
            col = col_map.get(item.type)
            if col:
                ws[f"{col}{current_row}"] = float(item.montant)
            current_row += 1

        total = float(note.montant_total)
        ws["J9"] = total
        ws["J10"] = f"{total} {note.devise}"
        ws["J12"] = f"{total} {note.devise}"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{note.reference}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class NoteDeFraisExportPdfView(APIView):
    """Export d'une note de frais en PDF"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            note = NoteDeFrais.objects.get(pk=pk)
        except NoteDeFrais.DoesNotExist:
            return Response({"detail": "Note de frais introuvable"}, status=404)

        template_path = os.path.join(settings.BASE_DIR, "templates", "NOTE-DE-DEPENSES-MRU.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active

        ws["J1"] = note.reference
        ws["J2"] = note.date_creation.date().strftime("%d/%m/%Y")

        start_row = 7
        current_row = start_row
        col_map = {
            "nourriture": "C",
            "hebergement": "D",
            "medicament": "E",
            "carburant": "F",
            "entretien": "G",
            "telecom": "H",
            "avance": "I",
            "divers": "J",
        }
        for item in note.items.all():
            ws[f"A{current_row}"] = item.libelle
            col = col_map.get(item.type)
            if col:
                ws[f"{col}{current_row}"] = float(item.montant)
            current_row += 1

        total = float(note.montant_total)
        ws["J9"] = total
        ws["J10"] = f"{total} {note.devise}"
        ws["J12"] = f"{total} {note.devise}"

        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "note.xlsx")
            pdf_path = os.path.join(tmpdir, "note.pdf")
            wb.save(xlsx_path)

            cmd = [
                "soffice",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                tmpdir,
                xlsx_path,
            ]
            subprocess.run(cmd, check=True)

            return FileResponse(
                open(pdf_path, "rb"),
                as_attachment=True,
                filename=f"{note.reference}.pdf",
                content_type="application/pdf",
            )


# ==================== DEVIS (Spec 4) ====================

class DevisListCreateView(generics.ListCreateAPIView):
    """Liste tous les devis ou crée un nouveau devis - Spec 4"""
    queryset = Devis.objects.all().order_by('-date_creation')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return DevisCreateSerializer
        elif self.request.method == 'GET':
            return DevisDetailSerializer
        return DevisSerializer


class DevisRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un devis - Spec 4"""
    queryset = Devis.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return DevisDetailSerializer
        if self.request.method in ['PUT', 'PATCH']:
            return DevisCreateSerializer
        return DevisSerializer


class DevisValiderView(APIView):
    """Valide ou rejette un devis et génère une facture si validé - Spec 4"""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        devis = get_object_or_404(Devis, pk=pk)
        nouveau_statut = request.data.get('status')

        if nouveau_statut not in ['valide', 'rejete']:
            return Response(
                {"error": "Statut invalide"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                # 1. Mise à jour du devis
                devis.status = nouveau_statut
                devis.valideur = request.user
                devis.date_validation = datetime.now()
                devis.save()

                facture = None
                if nouveau_statut == 'valide':
                    facture = Facture.objects.create(
                        client=devis.client,
                        port_arrive=devis.port_arrive,
                        vessel=devis.vessel,
                        voyage=devis.voyage,
                        eta=devis.eta,
                        etd=devis.etd,
                        bl=devis.bl,
                        type=devis.type,
                        description=devis.description,
                        is_excluding_customs=devis.is_excluding_customs,
                        volume=devis.volume,
                        poids=devis.poids,
                        commentaire=devis.commentaire,
                        tva=devis.tva,
                        devise=devis.devise,
                        createur=request.user,
                        status='attente',
                        est_privee=False
                    )

                    items_devis = devis.items.all()
                    items_facture = [
                        ItemFacture(
                            facture=facture,
                            libelle=item.libelle,
                            prix_unitaire=item.prix_unitaire,
                            quantite=item.quantite
                        ) for item in items_devis
                    ]
                    ItemFacture.objects.bulk_create(items_facture)

            message = f"Devis {nouveau_statut} avec succès"
            if facture:
                message += f". Facture {facture.reference} générée automatiquement."

            return Response({
                "message": message,
                "status": devis.status,
                "facture_reference": facture.reference if facture else None
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Une erreur est survenue lors de la validation : {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class DevisParClientView(generics.ListAPIView):
    """Liste les devis d'un client spécifique"""
    serializer_class = DevisSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        client_id = self.kwargs.get('client_id')
        return Devis.objects.filter(client_id=client_id).order_by('-date_creation')


class DevisAjouterItemView(APIView):
    """Ajoute un item à un devis existant"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            devis = Devis.objects.get(pk=pk)
        except Devis.DoesNotExist:
            return Response(
                {"detail": "Devis introuvable"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = ItemDevisSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(devis=devis)
            return Response(
                DevisDetailSerializer(devis).data,
                status=status.HTTP_201_CREATED
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DevisConvertirEnFactureView(APIView):
    """Convertit un devis en facture"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            devis = Devis.objects.get(pk=pk)
        except Devis.DoesNotExist:
            return Response(
                {"detail": "Devis introuvable"},
                status=status.HTTP_404_NOT_FOUND
            )

        with db_transaction.atomic():
            facture = Facture.objects.create(
                client=devis.client,
                port_arrive=devis.port_arrive,
                vessel=devis.vessel,
                voyage=devis.voyage,
                eta=devis.eta,
                etd=devis.etd,
                bl=devis.bl,
                tva=devis.tva,
                devise=devis.devise,
                createur=request.user
            )

            for item_devis in devis.items.all():
                ItemFacture.objects.create(
                    facture=facture,
                    libelle=item_devis.libelle,
                    prix_unitaire=item_devis.prix_unitaire,
                    quantite=item_devis.quantite
                )

        return Response(
            {
                "message": "Devis converti en facture avec succès",
                "facture": FactureDetailSerializer(facture).data
            },
            status=status.HTTP_201_CREATED
        )


# ==================== FACTURE (Spec 3) ====================

class FactureListCreateView(generics.ListCreateAPIView):
    """Liste toutes les factures ou crée une nouvelle - Spec 3"""
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return FactureCreateSerializer
        elif self.request.method == 'GET':
            return FactureDetailSerializer
        return FactureDetailSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = Facture.objects.all().order_by('-date_creation')

        if user.type != 'directeur_general':
            queryset = queryset.filter(est_privee=False)

        return queryset


class FactureRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime une facture - Spec 3"""
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return FactureDetailSerializer
        if self.request.method in ['PUT', 'PATCH']:
            return FactureCreateSerializer
        return FactureSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = Facture.objects.all()

        # Factures privées visibles uniquement par le DG (Spec 3)
        if user.type != 'directeur_general':
            queryset = queryset.filter(est_privee=False)

        return queryset


class FactureValiderView(APIView):
    """Valide ou rejette une facture - Spec 3"""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        # Validation par DG ou Comptable (Spec 3)
        if request.user.type not in ['directeur_general', 'comptable']:
            return Response(
                {"error": "Vous n'avez pas la permission de valider"},
                status=status.HTTP_403_FORBIDDEN
            )

        facture = get_object_or_404(Facture, pk=pk)
        nouveau_statut = request.data.get('status')

        if nouveau_statut not in ['valide', 'rejete']:
            return Response(
                {"error": "Statut invalide"},
                status=status.HTTP_400_BAD_REQUEST
            )

        facture.status = nouveau_statut
        facture.valideur = request.user
        facture.date_validation = datetime.now()
        facture.save()

        return Response({
            "message": f"Facture {nouveau_statut}e avec succès",
            "status": facture.status
        })


class FactureParClientView(generics.ListAPIView):
    """Liste les factures d'un client spécifique"""
    serializer_class = FactureSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        client_id = self.kwargs.get('client_id')
        user = self.request.user

        queryset = Facture.objects.filter(client_id=client_id).order_by('-date_creation')

        # Filtrer les factures privées
        if user.type != 'directeur_general':
            queryset = queryset.filter(est_privee=False)

        return queryset


class FactureAjouterItemView(APIView):
    """Ajoute un item à une facture existante"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            facture = Facture.objects.get(pk=pk)
        except Facture.DoesNotExist:
            return Response(
                {"detail": "Facture introuvable"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = ItemFactureSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(facture=facture)
            return Response(
                FactureDetailSerializer(facture).data,
                status=status.HTTP_201_CREATED
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ==================== BON DE COMMANDE (Spec 10) ====================

class BonCommandeListCreateView(generics.ListCreateAPIView):
    """Liste tous les bons de commande ou crée un nouveau - Spec 10"""
    queryset = BonCommande.objects.all().order_by('-date_creation')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return BonCommandeCreateSerializer
        elif self.request.method == 'GET':
            return BonCommandeDetailSerializer
        return BonCommandeSerializer


class BonCommandeRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un bon de commande - Spec 10"""
    queryset = BonCommande.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return BonCommandeDetailSerializer
        if self.request.method in ['PUT', 'PATCH']:
            return BonCommandeCreateSerializer
        return BonCommandeSerializer


class BonCommandeValiderView(APIView):
    """Valide ou rejette un bon de commande - Spec 10"""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        bon_commande = get_object_or_404(BonCommande, pk=pk)
        nouveau_statut = request.data.get('status')

        if nouveau_statut not in ['valide', 'rejete']:
            return Response(
                {"error": "Statut invalide"},
                status=status.HTTP_400_BAD_REQUEST
            )

        bon_commande.status = nouveau_statut
        bon_commande.valideur = request.user
        bon_commande.date_validation = datetime.now()
        bon_commande.save()

        return Response({
            "message": f"Bon de commande {nouveau_statut} avec succès",
            "status": bon_commande.status
        })


class BonCommandeAjouterItemView(APIView):
    """Ajoute un item à un bon de commande existant - Spec 10"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            bon_commande = BonCommande.objects.get(pk=pk)
        except BonCommande.DoesNotExist:
            return Response(
                {"detail": "Bon de commande introuvable"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = ItemBonCommandeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(bon_commande=bon_commande)
            return Response(
                BonCommandeDetailSerializer(bon_commande).data,
                status=status.HTTP_201_CREATED
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ==================== UTILISATEUR ====================

class UtilisateurListCreateView(generics.ListCreateAPIView):
    """Liste tous les utilisateurs ou crée un nouvel utilisateur"""
    queryset = Utilisateur.objects.all().order_by('nom', 'prenom')
    serializer_class = UtilisateurSerializer
    permission_classes = [IsAuthenticated]


class UtilisateurRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un utilisateur spécifique"""
    queryset = Utilisateur.objects.all()
    serializer_class = UtilisateurCustomSerializer
    permission_classes = [IsAuthenticated]


class UtilisateurParTypeView(generics.ListAPIView):
    """Liste les utilisateurs filtrés par type"""
    serializer_class = UtilisateurSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        type_utilisateur = self.request.query_params.get('type')
        queryset = Utilisateur.objects.all().order_by('nom', 'prenom')

        if type_utilisateur:
            queryset = queryset.filter(type=type_utilisateur)

        return queryset


# ==================== AUTHENTIFICATION ====================

class SeConnecter(TokenObtainPairView):
    """Connexion de l'utilisateur avec username et mot de passe"""
    serializer_class = ConnexionSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        username = serializer.validated_data['username']
        mot_de_passe = serializer.validated_data['mot_de_passe']

        utilisateur = Utilisateur.objects.filter(username=username).first()

        if utilisateur:
            utilisateur = authenticate(request, username=username, password=mot_de_passe)
            if utilisateur:
                refresh = RefreshToken.for_user(utilisateur)

                utilisateur_data = UtilisateurSerializer(utilisateur).data

                data = {
                    'token': str(refresh.access_token),
                    'refresh': str(refresh),
                    'utilisateur': utilisateur_data,
                }

                return Response(data, status=status.HTTP_200_OK)
            else:
                return Response(
                    {"detail": "Nom d'utilisateur ou mot de passe incorrect"},
                    status=status.HTTP_401_UNAUTHORIZED
                )
        else:
            return Response(
                {"detail": "Utilisateur introuvable"},
                status=status.HTTP_404_NOT_FOUND
            )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def utilisateur_profil(request):
    """Récupère le profil de l'utilisateur connecté"""
    utilisateur = request.user
    try:
        utilisateur_data = UtilisateurSerializer(utilisateur)
        return JsonResponse(utilisateur_data.data)
    except Exception as e:
        return JsonResponse({
            "status": 500,
            "message": f"Échec du chargement du profil utilisateur: {e}"
        }, status=500)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def modifier_informations(request):
    """Modifie les informations de l'utilisateur connecté"""
    user = request.user
    serializer = UtilisateurSerializer(user, data=request.data, partial=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def modifier_mot_de_passe(request):
    """Modifie le mot de passe de l'utilisateur connecté"""
    utilisateur = request.user
    try:
        data = json.loads(request.body.decode('utf-8'))
        ancien = data.get('ancien', '')
        nouveau = data.get('nouveau', '')

        if not ancien or not nouveau:
            return Response(
                {"message": "L'ancien et le nouveau mot de passe sont requis"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if check_password(ancien, utilisateur.password):
            utilisateur.password = make_password(nouveau)
            utilisateur.save()

            return Response(
                {"message": "Mot de passe modifié avec succès"},
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                {"message": "Ancien mot de passe incorrect"},
                status=status.HTTP_400_BAD_REQUEST
            )

    except Exception as e:
        return Response(
            {"message": f"Une erreur est survenue: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==================== STATISTIQUES ====================

class StatistiquesGeneralesView(APIView):
    """Retourne les statistiques générales du système"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        stats = {
            'total_rotations': Rotation.objects.count(),
            'rotations_entree': Rotation.objects.filter(type='entree').count(),
            'rotations_sortie': Rotation.objects.filter(type='sortie').count(),
            'total_rotations_entrantes': RotationEntrante.objects.count(),
            'total_rotations_sortantes': RotationSortante.objects.count(),
            'total_clients': Client.objects.count(),
            'total_fournisseurs': Fournisseur.objects.count(),
            'total_types_materiel': TypeMateriel.objects.count(),
            'total_expressions_besoin': ExpressionBesoin.objects.count(),
            'total_devis': Devis.objects.count(),
            'total_factures': Facture.objects.count(),
            'total_notes_frais': NoteDeFrais.objects.count(),
            'total_bons_commande': BonCommande.objects.count(),
            'total_utilisateurs': Utilisateur.objects.count(),
        }

        # Montants par devise (Devis)
        devis_par_devise = {}
        for devise_code, devise_nom in Devis.DEVISES:
            devis = Devis.objects.filter(devise=devise_code)
            total = sum(d.montant_total for d in devis)
            devis_par_devise[devise_code] = {
                'nom': devise_nom,
                'total': float(total),
                'nombre': devis.count()
            }

        # Montants par devise (Factures)
        factures_par_devise = {}
        for devise_code, devise_nom in Facture.DEVISES:
            factures = Facture.objects.filter(devise=devise_code)
            total = sum(f.montant_total for f in factures)
            factures_par_devise[devise_code] = {
                'nom': devise_nom,
                'total': float(total),
                'nombre': factures.count()
            }

        stats['devis_par_devise'] = devis_par_devise
        stats['factures_par_devise'] = factures_par_devise

        return Response(stats, status=status.HTTP_200_OK)


class StatistiquesClientView(APIView):
    """Retourne les statistiques d'un client spécifique"""
    permission_classes = [IsAuthenticated]

    def get(self, request, client_id):
        try:
            client = Client.objects.get(pk=client_id)
        except Client.DoesNotExist:
            return Response(
                {"detail": "Client introuvable"},
                status=status.HTTP_404_NOT_FOUND
            )

        devis = Devis.objects.filter(client=client)
        factures = Facture.objects.filter(client=client)

        stats = {
            'client': ClientSerializer(client).data,
            'total_devis': devis.count(),
            'total_factures': factures.count(),
            'montant_total_devis': sum(d.montant_total for d in devis),
            'montant_total_factures': sum(f.montant_total for f in factures),
        }

        return Response(stats, status=status.HTTP_200_OK)





# ==================== BON À DÉLIVRER (BAD) ====================

class BADListCreateView(generics.ListCreateAPIView):
    """Liste tous les BAD ou crée un nouveau BAD"""
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return BADCreateSerializer
        elif self.request.method == 'GET':
            return BADSerializer
        return BADSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = BAD.objects.all().order_by('-date_creation')

        # Filtre optionnel par client via query params
        client_id = self.request.query_params.get('client')
        if client_id:
            queryset = queryset.filter(client_id=client_id)

        return queryset


class BADRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un BAD"""
    queryset = BAD.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return BADCreateSerializer
        return BADSerializer


class BADValiderItemView(APIView):
    """Permet à un agent ou directeur de valider un item spécifique d'un BAD"""
    permission_classes = [IsAuthenticated]

    def patch(self, request, item_id):
        item = get_object_or_404(ItemBAD, pk=item_id)

        # Seuls certains rôles peuvent valider les items logistiques
        if request.user.type not in ['agent_port', 'directeur_operations', 'directeur_general']:
            return Response(
                {"error": "Vous n'avez pas la permission de valider cet item"},
                status=status.HTTP_403_FORBIDDEN
            )

        item.valideur = request.user
        item.save()

        return Response({
            "message": "Item validé avec succès",
            "valideur": request.user.get_full_name()
        })


class BADParFactureView(generics.ListAPIView):
    """Récupère les BAD associés à une facture spécifique"""
    serializer_class = BADSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        facture_id = self.kwargs.get('facture_id')
        return BAD.objects.filter(facture_id=facture_id)


class BADExportPdfView(APIView):
    """
    Vue placeholder pour l'export PDF du BAD
    Suit la logique de vos autres exports (soffice/template)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        bad = get_object_or_404(BAD, pk=pk)

        # Ici vous pourriez utiliser un template Excel spécifique au BAD
        # similaire à votre NoteDeFraisExportPdfView

        return Response({"message": "Fonctionnalité d'export en cours de déploiement pour le format BAD"})




# ==================== ARCHIVES DOCUMENTAIRES (GED) ====================

class DocumentArchiveListCreateView(generics.ListCreateAPIView):
    """Liste tous les documents archivés ou upload un nouveau document"""
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return DocumentArchiveCreateSerializer
        return DocumentArchiveSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = DocumentArchive.objects.all().order_by('-date_upload')

        # Filtrage par type si spécifié dans l'URL (?type_doc=BL)
        type_doc = self.request.query_params.get('type_doc')
        if type_doc:
            queryset = queryset.filter(type_doc=type_doc)

        # Optionnel : Si vous voulez que les assistants ne voient que leurs propres uploads
        # if user.type == 'assistant':
        #     queryset = queryset.filter(cree_par=user)

        return queryset


class DocumentArchiveRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un document archivé"""
    queryset = DocumentArchive.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return DocumentArchiveCreateSerializer
        return DocumentArchiveSerializer


class DocumentArchiveRechercheView(generics.ListAPIView):
    """Recherche de documents par titre ou description"""
    serializer_class = DocumentArchiveSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        query = self.request.query_params.get('q', '')
        return DocumentArchive.objects.filter(
            Q(titre__icontains=query) |
            Q(description__icontains=query)
        ).order_by('-date_upload')


class RapportJournalierStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        target_date = request.query_params.get('date')
        mouvement = request.query_params.get('type_mouvement', 'sorties')

        if not target_date:
            return Response({"error": "Date manquante"}, status=400)

        # 1. Filtrage de base
        if mouvement == 'entrees':
            qs = RotationEntrante.objects.filter(date_arrivee__date=target_date)
        else:
            qs = RotationSortante.objects.filter(date_sortie__date=target_date)

        qs = qs.select_related('client', 'type_materiel')

        # 2. Récapitulatif Global (Déjà correct dans votre code, on le garde)
        stats_par_type = qs.values('type_materiel__nom').annotate(total=Sum('quantite'))
        recapitulatif = {item['type_materiel__nom']: item['total'] for item in stats_par_type}

        # 3. Structure par client avec REGROUPEMENT (Addition)
        # On utilise un dictionnaire : { "Nom Client": { "Type Materiel": Quantité_Totale } }
        clients_map = {}

        for item in qs:
            c_nom = item.client.nom
            t_nom = item.type_materiel.nom
            qte = item.quantite or 0

            if c_nom not in clients_map:
                clients_map[c_nom] = {}

            # Si le type de matériel existe déjà pour ce client, on additionne
            if t_nom in clients_map[c_nom]:
                clients_map[c_nom][t_nom] += qte
            else:
                clients_map[c_nom][t_nom] = qte

        # 4. Formatage de la réponse pour le frontend
        detailed_stats = []
        for client_nom, materiels in clients_map.items():
            # Transformer le sous-dictionnaire en liste de mouvements groupés
            mouvements_groupes = [
                {"type": t_nom, "quantite": qte_totale}
                for t_nom, qte_totale in materiels.items()
            ]
            detailed_stats.append({
                "client": client_nom,
                "mouvements": mouvements_groupes
            })

        return Response({
            "date_consultee": target_date,
            "type_mouvement": mouvement,
            "recapitulatif": recapitulatif,
            "details_par_client": detailed_stats
        })


# --- Statistiques Globales ---
class StatistiquesGlobalesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        mouvement = request.query_params.get('type_mouvement', 'sorties')

        if mouvement == 'entrees':
            queryset = RotationEntrante.objects.filter(date_arrivee__date__range=[start_date, end_date]).annotate(jour=TruncDate('date_arrivee'))
        else:
            queryset = RotationSortante.objects.filter(date_sortie__date__range=[start_date, end_date]).annotate(jour=TruncDate('date_sortie'))

        tous_les_clients = list(Client.objects.values_list('nom', flat=True))
        matrice = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
        totaux_colonnes = defaultdict(lambda: defaultdict(int))

        for s in queryset.select_related('client', 'type_materiel'):
            jour_str = str(s.jour)
            client_nom = s.client.nom
            type_nom = s.type_materiel.nom
            matrice[jour_str][client_nom][type_nom] += s.quantite
            totaux_colonnes[client_nom][type_nom] += s.quantite

        lignes_tableau = []
        for index, date_cle in enumerate(sorted(matrice.keys()), start=1):
            ligne = {"label": f"DAY {index}", "date": date_cle, "clients": {}}
            for client in tous_les_clients:
                mvts = matrice[date_cle].get(client, {})
                ligne["clients"][client] = "\n".join([f"{qty} {tp.upper()}" for tp, qty in mvts.items()])
            lignes_tableau.append(ligne)

        total_final = {"label": "TOTAL", "clients": {}}
        for client in tous_les_clients:
            mvts_totaux = totaux_colonnes.get(client, {})
            total_final["clients"][client] = "\n".join([f"{qty} {tp.upper()}" for tp, qty in mvts_totaux.items()])

        return Response({
            "type_mouvement": mouvement,
            "colonnes": tous_les_clients,
            "lignes": lignes_tableau,
            "total": total_final,
            "mouvementType": mouvement
        })
    

class PDAListCreateView(generics.ListCreateAPIView):
    """Liste tous les PDA ou crée un nouveau PDA"""
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return PDACreateUpdateSerializer
        return PDASerializer

    def get_queryset(self):
        # Tri par date de création décroissante
        queryset = PDA.objects.all().order_by('-date')

        # Recherche par nom du client (lié via ForeignKey)
        client_name = self.request.query_params.get('client')
        if client_name:
            queryset = queryset.filter(client__nom__icontains=client_name)

        # Filtre par navire
        vessel = self.request.query_params.get('vessel')
        if vessel:
            queryset = queryset.filter(vessel_name__icontains=vessel)

        return queryset

    def perform_create(self, serializer):
        """
        Cette méthode intercepte la sauvegarde pour ajouter l'utilisateur actuel.
        Assure-toi que ton modèle PDA possède : createur = models.ForeignKey(User, ...)
        """
        # Si ton modèle n'a PAS de champ 'createur', retire simplement cet argument.
        # Si tu veux enregistrer qui a fait le PDA :
        serializer.save(createur=self.request.user)


class PDARetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    """Récupère, modifie ou supprime un PDA"""
    queryset = PDA.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return PDACreateUpdateSerializer
        return PDASerializer